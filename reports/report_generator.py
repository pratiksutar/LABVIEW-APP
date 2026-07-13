"""
Report Generator
Exports workstation status, utilization reports, and audit logs
to CSV and Excel (openpyxl) formats.
"""
from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from domain.models.workstation import Workstation

REPORTS_DIR = Path.home() / ".labview" / "reports"

# ─── Styles ──────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1A1A2E") if OPENPYXL_AVAILABLE else None
HDR_FONT  = Font(color="A78BFA", bold=True, name="Segoe UI", size=10) if OPENPYXL_AVAILABLE else None
BODY_FONT = Font(color="E2E8F0", name="Segoe UI", size=10)             if OPENPYXL_AVAILABLE else None
ALT_FILL  = PatternFill("solid", fgColor="22223A")                     if OPENPYXL_AVAILABLE else None


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


class ReportGenerator:
    """Generates utilization, status, and audit log reports."""

    def __init__(self) -> None:
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # ─── Workstation status CSV ───────────────────────────────────
    def export_status_csv(self, workstations: List[Workstation]) -> Path:
        path = REPORTS_DIR / f"labview_status_{_ts()}.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Name", "Area", "Status",
                              "Power State", "Watts", "Device ID",
                              "Maintenance", "Last Updated"])
            for ws in workstations:
                writer.writerow([
                    ws.id, ws.name, ws.area, ws.status.display_name,
                    "ON" if ws.power_state else "OFF",
                    f"{ws.power_consumption:.1f}", ws.device_id or "",
                    "Yes" if ws.is_maintenance else "No",
                    ws.last_updated.strftime("%Y-%m-%d %H:%M:%S") if ws.last_updated else "",
                ])
        logger.info(f"[Reports] CSV exported → {path}")
        return path

    # ─── Audit log CSV ────────────────────────────────────────────
    def export_audit_csv(self, rows: list) -> Path:
        path = REPORTS_DIR / f"labview_audit_{_ts()}.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Timestamp", "User", "Action", "Entity",
                              "Entity ID", "Old Value", "New Value"])
            for r in rows:
                ts = r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else ""
                writer.writerow([ts, r.username, r.action, r.entity,
                                  r.entity_id or "", r.old_value or "",
                                  r.new_value or ""])
        logger.info(f"[Reports] Audit CSV exported → {path}")
        return path

    # ─── Utilization report CSV ──────────────────────────────────
    def export_utilization_csv(self, report_data: Dict[str, Any]) -> Path:
        label  = report_data.get("label", "report").replace(" ", "_").replace(",", "")
        path   = REPORTS_DIR / f"labview_utilization_{label}_{_ts()}.csv"
        stats  = report_data.get("stats", {})
        ws_rows = report_data.get("per_workstation", [])

        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["LabView Utilization Report"])
            writer.writerow(["Period", report_data.get("label", "")])
            writer.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
            writer.writerow([])

            writer.writerow(["Summary"])
            writer.writerow(["In Use %", stats.get("in_use_pct", 0)])
            writer.writerow(["Idle %",   stats.get("idle_pct", 0)])
            writer.writerow(["Available %", stats.get("available_pct", 0)])
            writer.writerow(["Avg Power (W)", stats.get("avg_power_w", 0)])
            writer.writerow(["Total Records", stats.get("total_records", 0)])
            writer.writerow([])

            if ws_rows:
                writer.writerow(["Workstation", "In Use %", "Idle %", "Available %", "Records"])
                for ws in ws_rows:
                    writer.writerow([ws["name"], ws["in_use_pct"],
                                     ws["idle_pct"], ws["avail_pct"], ws["total"]])

        logger.info(f"[Reports] Utilization CSV → {path}")
        return path

    # ─── Workstation status Excel ─────────────────────────────────
    def export_status_xlsx(self, workstations: List[Workstation]) -> Path:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed: pip install openpyxl")

        path = REPORTS_DIR / f"labview_status_{_ts()}.xlsx"
        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = "Workstation Status"

        headers = ["ID", "Name", "Area", "Status",
                   "Power", "Watts", "Device ID", "Maintenance", "Last Updated"]
        self._write_header_row(ws, headers)

        status_colors = {
            "available":    "10B981", "idle":         "F59E0B",
            "in_use":       "3B82F6", "disconnected": "6B7280",
            "maintenance":  "F97316", "failure":      "EF4444",
        }
        for row_idx, w in enumerate(workstations, start=2):
            data = [w.id, w.name, w.area or "", w.status.display_name,
                    "ON" if w.power_state else "OFF",
                    round(w.power_consumption, 1), w.device_id or "",
                    "Yes" if w.is_maintenance else "No",
                    w.last_updated.strftime("%Y-%m-%d %H:%M:%S") if w.last_updated else ""]
            self._write_data_row(ws, row_idx, data)
            # Color the Status cell
            cell = ws.cell(row=row_idx, column=4)
            color = status_colors.get(w.status.value, "6B7280")
            cell.font = Font(color=color, bold=True, name="Segoe UI", size=10)

        col_widths = [6, 24, 16, 16, 8, 10, 32, 14, 22]
        for col, w_ in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w_

        wb.save(path)
        logger.info(f"[Reports] Status Excel → {path}")
        return path

    # ─── Utilization Excel ────────────────────────────────────────
    def export_utilization_xlsx(self, report_data: Dict[str, Any]) -> Path:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed: pip install openpyxl")

        label   = report_data.get("label", "report").replace(" ", "_").replace(",", "")
        path    = REPORTS_DIR / f"labview_utilization_{label}_{_ts()}.xlsx"
        stats   = report_data.get("stats", {})
        ws_rows = report_data.get("per_workstation", [])

        wb = openpyxl.Workbook()

        # ── Summary sheet ──
        ws_sum = wb.active
        ws_sum.title = "Summary"

        title_font = Font(color="A78BFA", bold=True, name="Segoe UI", size=14)
        ws_sum["A1"] = "LabView Utilization Report"
        ws_sum["A1"].font = title_font
        ws_sum["A2"] = report_data.get("label", "")
        ws_sum["A2"].font = Font(color="9CA3AF", name="Segoe UI", size=11)

        summary_rows = [
            ["In Use %",    stats.get("in_use_pct", 0)],
            ["Idle %",      stats.get("idle_pct", 0)],
            ["Available %", stats.get("available_pct", 0)],
            ["Maintenance %", stats.get("maintenance_pct", 0)],
            ["Avg Power (W)", stats.get("avg_power_w", 0)],
            ["Total Records", stats.get("total_records", 0)],
        ]
        for i, (k, v) in enumerate(summary_rows, start=4):
            ws_sum.cell(row=i, column=1, value=k).font = Font(
                color="9CA3AF", name="Segoe UI", size=10)
            ws_sum.cell(row=i, column=2, value=v).font = Font(
                color="E2E8F0", bold=True, name="Segoe UI", size=10)
        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 14

        # ── Per-workstation sheet ──
        if ws_rows:
            ws_detail = wb.create_sheet("By Workstation")
            self._write_header_row(ws_detail,
                                   ["Workstation", "In Use %", "Idle %",
                                    "Available %", "Total Records"])
            for i, row in enumerate(ws_rows, start=2):
                self._write_data_row(ws_detail, i,
                                     [row["name"], row["in_use_pct"],
                                      row["idle_pct"], row["avail_pct"],
                                      row["total"]])
            ws_detail.column_dimensions["A"].width = 28
            for col in ["B", "C", "D", "E"]:
                ws_detail.column_dimensions[col].width = 14

        # ── Chart data sheet ──
        ws_chart = wb.create_sheet("Chart Data")
        labels = report_data.get("chart_labels", [])
        values = report_data.get("chart_values", [])
        ws_chart.cell(row=1, column=1, value="Label").font = HDR_FONT
        ws_chart.cell(row=1, column=2, value="In-Use Count").font = HDR_FONT
        for i, (lbl, val) in enumerate(zip(labels, values), start=2):
            ws_chart.cell(row=i, column=1, value=lbl)
            ws_chart.cell(row=i, column=2, value=val)
        ws_chart.column_dimensions["A"].width = 14
        ws_chart.column_dimensions["B"].width = 14

        wb.save(path)
        logger.info(f"[Reports] Utilization Excel → {path}")
        return path

    # ─── Audit log Excel ──────────────────────────────────────────
    def export_audit_xlsx(self, rows: list) -> Path:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl not installed: pip install openpyxl")

        path = REPORTS_DIR / f"labview_audit_{_ts()}.xlsx"
        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = "Audit Log"

        headers = ["Timestamp", "User", "Action", "Entity",
                   "Entity ID", "Old Value", "New Value"]
        self._write_header_row(ws, headers)

        for row_idx, r in enumerate(rows, start=2):
            ts = r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else ""
            self._write_data_row(ws, row_idx,
                                 [ts, r.username, r.action, r.entity,
                                  r.entity_id or "", r.old_value or "",
                                  r.new_value or ""])

        col_widths = [20, 14, 26, 18, 10, 30, 30]
        for col, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

        wb.save(path)
        logger.info(f"[Reports] Audit Excel → {path}")
        return path

    # ─── Shared helpers ──────────────────────────────────────────
    def _write_header_row(self, ws, headers: List[str]) -> None:
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    def _write_data_row(self, ws, row_idx: int, values: list) -> None:
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18
